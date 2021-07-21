#!/usr/bin/env python
# coding: utf-8

# In[55]:


import xlrd
from xlrd import open_workbook


# In[56]:


loc = ("excel.xlsx")  


# In[60]:


wb = xlrd.open_workbook(loc)     
sheet = wb.sheet_by_index(0)
sheet.cell_value(5,4)  


# In[62]:


for i in range(0,sheet.nrows):
    for j in range(0,sheet.ncols):
        print(sheet.cell_value(i, j))
  


# In[70]:


# Reading and writing in excel can be done by single module 
import openpyxl 
from openpyxl.utils.cell import get_column_letter 

workbook = openpyxl.load_workbook('excel.xlsx') 
workbook.sheetnames 
worksheet = workbook["Sheet1"]


# In[71]:


# Number of rows 
number_of_rows = worksheet.max_row 
  
# Number of columns 
number_of_columns = worksheet.max_column


# In[72]:


word=input('Enter the word you want to replace :')
new_word=input("Enter the replaced word : ")
replacementTextKeyPairs ={'{}'.format(word):'{}'.format(new_word)}
#print(replacementTextKeyPairs)


# In[73]:


for i in range(number_of_columns): 
    for k in range(number_of_rows): 
          
        cellValue = str(worksheet[get_column_letter(i+1)+str(k+1)].value) 
          
        for key in replacementTextKeyPairs.keys(): 
              
            if str(cellValue) == key: 
                newCellValue = replacementTextKeyPairs.get(key) 
                worksheet[get_column_letter(i+1)+str(k+1)] = str(newCellValue) 


# In[74]:


workbook.save('sampleexcelwithreplacedtextusingopenpyxl.xlsx')


# In[75]:


loc1 = ("sampleexcelwithreplacedtextusingopenpyxl.xlsx")
wb1 = xlrd.open_workbook(loc1)     
sheet1 = wb1.sheet_by_index(0)
for i in range(0,sheet1.nrows):
    for j in range(0,sheet1.ncols):
        print(sheet1.cell_value(i, j))
  


# In[83]:


word_search=input('Enter the word you want to search :')
count=0
for i in range(0,sheet.nrows):
    for j in range(0,sheet.ncols):
        if sheet.cell_value(i,j) == word_search:
            print("Word Found at the position ",i,'*',j)
            count=count+1
if count==0:
    print("Word is not Found")


# In[ ]:




